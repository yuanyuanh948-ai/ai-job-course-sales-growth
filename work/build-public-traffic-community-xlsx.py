from __future__ import annotations

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
SOURCE_MD = ROOT / "docs" / "public-traffic-community-operation-plan.md"
OUTPUT_FILE = ROOT / "outputs" / "AI就业班公域社群运营方案_完整版_20260623.xlsx"


def section_between(text: str, start_pattern: str, end_pattern: str | None = None) -> str:
    start = re.search(start_pattern, text, flags=re.M)
    if not start:
        return ""
    if end_pattern:
        end = re.search(end_pattern, text[start.end() :], flags=re.M)
        if end:
            return text[start.end() : start.end() + end.start()].strip()
    return text[start.end() :].strip()


def clean_markdown_block(raw: str) -> str:
    lines: list[str] = []
    for line in raw.strip().splitlines():
        line = line.rstrip()
        if line.startswith(">"):
            line = line[1:].lstrip()
        if line.strip() in {"```", "```text"}:
            continue
        lines.append(line)
    return "\n".join(lines).strip()


def parse_markdown_table(raw: str) -> list[list[str]]:
    rows: list[list[str]] = []
    for line in raw.splitlines():
        line = line.strip()
        if not line.startswith("|") or not line.endswith("|"):
            continue
        cells = [cell.strip() for cell in line.strip("|").split("|")]
        if cells and all(set(cell) <= {"-", " ", ":"} for cell in cells):
            continue
        rows.append(cells)
    return rows


def extract_labeled(block: str, label: str, labels: list[str]) -> str:
    match = re.search(re.escape(label), block)
    if not match:
        return ""
    rest = block[match.end() :]
    next_positions = [
        pos
        for other in labels
        if other != label
        for pos in [rest.find(other)]
        if pos >= 0
    ]
    if next_positions:
        rest = rest[: min(next_positions)]
    return clean_markdown_block(rest)


def parse_daily_posts(text: str) -> list[list[str]]:
    daily = section_between(text, r"^## 6\. 14 天详细内容排期表$", r"^## 7\.")
    labels = ["投送内容：", "互动问题：", "小号追问：", "老师号可回应：", "CTA："]
    rows = [["周次", "星期", "栏目", "主题", "投送内容", "互动问题", "小号追问", "老师号可回应", "CTA"]]
    for raw_block in re.split(r"(?m)^### ", daily):
        block = raw_block.strip()
        if not block.startswith("第 "):
            continue
        title, _, body = block.partition("\n")
        title_match = re.match(r"(第 \d+ 周)(周[一二三四五六日])：(.+)", title.strip())
        week, weekday, column = title_match.groups() if title_match else ("", "", title.strip())
        topic_match = re.search(r"主题：(.+)", body)
        rows.append(
            [
                week,
                weekday,
                column,
                topic_match.group(1).strip() if topic_match else "",
                extract_labeled(body, "投送内容：", labels),
                extract_labeled(body, "互动问题：", labels),
                extract_labeled(body, "小号追问：", labels),
                extract_labeled(body, "老师号可回应：", labels),
                extract_labeled(body, "CTA：", labels),
            ]
        )
    return rows


def parse_trigger_replies(text: str) -> list[list[str]]:
    section = section_between(text, r"^## 7\. 触发词自动回复模板$", r"^## 8\.")
    rows = [["触发词", "自动回复模板"]]
    for raw_block in re.split(r"(?m)^### ", section):
        block = raw_block.strip()
        if not block:
            continue
        title, _, body = block.partition("\n")
        keyword_match = re.search(r"`([^`]+)`", title)
        rows.append([keyword_match.group(1) if keyword_match else title, clean_markdown_block(body)])
    return rows


def parse_subsections(text: str, start: str, end: str | None, headers: list[str]) -> list[list[str]]:
    section = section_between(text, start, end)
    rows = [headers]
    for raw_block in re.split(r"(?m)^### ", section):
        block = raw_block.strip()
        if not block:
            continue
        title, _, body = block.partition("\n")
        rows.append([title.strip(), clean_markdown_block(body)])
    return rows


def simple_bullet_sheet(raw: str, header: str) -> list[list[str]]:
    lines = [line.strip() for line in raw.splitlines() if line.strip()]
    return [[header], *[[line] for line in lines]]


def add_sheet(wb: Workbook, title: str, rows: list[list[str]]) -> None:
    ws = wb.create_sheet(title)
    for row in rows:
        ws.append(row)
    style_sheet(ws)


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    if ws.max_row >= 1:
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, max((len(part) for part in value.splitlines()), default=0))
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 55)

    ws.freeze_panes = "A2"


def main() -> None:
    text = SOURCE_MD.read_text(encoding="utf-8")
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    add_sheet(
        wb,
        "00_使用说明",
        [
            ["项目", "内容"],
            ["源文件", str(SOURCE_MD.relative_to(ROOT))],
            ["输出文件", str(OUTPUT_FILE.relative_to(ROOT))],
            ["说明", "本工作簿由 Markdown 社群方案自动结构化生成，覆盖完整运营方案、排期和话术。"],
        ],
    )
    add_sheet(
        wb,
        "01_方案总览",
        [
            ["模块", "内容"],
            ["社群定位", clean_markdown_block(section_between(text, r"^## 1\. 社群定位$", r"^## 2\."))],
            ["用户路径", clean_markdown_block(section_between(text, r"^## 2\. 用户路径$", r"^## 3\."))],
            ["两周滚动原则", clean_markdown_block(section_between(text, r"^## 5\. 两周滚动内容节奏$", r"^### 5\.1"))],
        ],
    )
    add_sheet(wb, "02_角色分工", parse_subsections(text, r"^## 3\. 角色分工$", r"^## 4\.", ["角色", "职责与说明"]))
    add_sheet(wb, "03_每日SOP", parse_markdown_table(section_between(text, r"^### 4\.1 当日流程$", r"^### 4\.2")))

    structure = clean_markdown_block(section_between(text, r"^### 4\.2 每条内容必须包含的 5 个结构$", r"^### 4\.3"))
    ctas = clean_markdown_block(section_between(text, r"^### 4\.3 每条内容必须带一个轻 CTA$", r"^## 5\."))
    add_sheet(
        wb,
        "04_内容结构CTA",
        [["类型", "内容"]]
        + [["内容结构", line] for line in structure.splitlines() if line.strip()]
        + [["轻CTA", line] for line in ctas.splitlines() if line.strip()],
    )

    add_sheet(wb, "05_两周总排期", parse_markdown_table(section_between(text, r"^### 5\.1 两周总排期$", r"^## 6\.")))
    add_sheet(wb, "06_14天投送话术", parse_daily_posts(text))
    add_sheet(wb, "07_触发词回复", parse_trigger_replies(text))
    add_sheet(wb, "08_私信分层", parse_subsections(text, r"^## 8\. 私信分层话术$", r"^## 9\.", ["用户类型", "识别信号与话术"]))
    add_sheet(wb, "09_公开课预热", parse_subsections(text, r"^## 9\. 公开课预热链路$", r"^## 10\.", ["阶段", "话术与动作"]))
    add_sheet(wb, "10_诊断名额机制", parse_subsections(text, r"^## 10\. 周六诊断名额机制$", r"^## 11\.", ["模块", "规则"]))
    add_sheet(wb, "11_数据看板", parse_markdown_table(section_between(text, r"^## 11\. 数据看板$", r"^## 12\.")))
    add_sheet(wb, "12_表达边界", simple_bullet_sheet(section_between(text, r"^## 12\. 表达边界$", r"^## 13\."), "表达规则"))
    add_sheet(wb, "13_助手执行清单", simple_bullet_sheet(section_between(text, r"^## 13\. 助手执行清单$", r"^## 14\."), "执行项"))
    add_sheet(wb, "14_老师执行清单", simple_bullet_sheet(section_between(text, r"^## 14\. 内容交流老师执行清单$"), "执行项"))

    wb.save(OUTPUT_FILE)
    print(OUTPUT_FILE)


if __name__ == "__main__":
    main()
